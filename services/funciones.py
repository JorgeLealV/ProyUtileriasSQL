# -*- coding: utf-8 -*-
"""
Este módulo contiene un conjunto de funciones de utilidad para interactuar
entre una base de datos PostgreSQL y archivos de Excel o SQL.

Funciones:
- postgres_a_excel: Exporta una tabla de PostgreSQL a un archivo de Excel con formato.
- excel_to_postgres_inserts: Convierte una hoja de Excel en una serie de sentencias SQL INSERT.
- execute_sql_from_file: Ejecuta un script SQL en una base de datos PostgreSQL.
"""

# --- Importación de Librerías ---
# pandas: Es una librería fundamental para la manipulación y análisis de datos.
#         La usamos para leer de la base de datos y escribir en Excel de forma muy eficiente.
#         Su estructura de datos principal es el DataFrame, que es como una tabla en memoria.
import pandas as pd

# psycopg2: Es el adaptador de base de datos más popular para PostgreSQL en Python.
#           Permite conectar Python con la base de datos para ejecutar consultas.
import psycopg2

# os: Proporciona funciones para interactuar con el sistema operativo, como por ejemplo,
#     comprobar si un archivo ya existe en el disco.
import os

# openpyxl.styles: Es parte de la librería openpyxl, que pandas usa internamente para
#                  trabajar con archivos .xlsx. Importamos PatternFill y Font para poder
#                  darle estilo a las celdas de Excel (color de fondo y de letra).
from openpyxl.styles import PatternFill, Font


def postgres_a_excel(servidor, puerto, base_de_datos, usuario, password, tabla, archivo_excel, acumular_si_existe):
    """
    Se conecta a una BD PostgreSQL, lee una tabla y la guarda en una hoja de Excel.

    Esta función es útil para crear reportes o copias de seguridad de tablas
    en un formato legible y fácil de compartir como es Excel.
    """
    # --- 1. Preparación de la Conexión ---
    # Se construye una "cadena de conexión". Es un string con un formato especial
    # que le dice a psycopg2 todos los detalles para conectarse a la base de datos.
    conn_string = f"dbname='{base_de_datos}' user='{usuario}' host='{servidor}' password='{password}' port='{puerto}'"
    
    # Se inicializa la variable de conexión a None. Es una buena práctica para asegurar
    # que la variable exista, incluso si la conexión falla, y poder comprobarlo al final.
    conn = None

    # --- 2. Bloque Try...Except...Finally: Manejo Seguro de la Conexión ---
    # Este bloque es crucial para manejar errores y asegurar que, pase lo que pase,
    # la conexión a la base de datos se cierre correctamente, evitando dejar conexiones
    # abiertas que consuman recursos.
    try:
        # --- 3. Conexión y Lectura de Datos ---
        # Se establece la conexión real con la base de datos.
        conn = psycopg2.connect(conn_string)
        print("Conexión a la base de datos establecida exitosamente.")

        # Se define la consulta SQL. Usamos f-string para insertar el nombre de la tabla.
        # Las comillas dobles alrededor de "{tabla}" son por si el nombre de la tabla
        # en PostgreSQL contiene mayúsculas o caracteres especiales.
        query = f'SELECT * FROM "{tabla}";'

        # Aquí ocurre la magia de pandas: pd.read_sql_query ejecuta la consulta
        # usando la conexión activa y carga TODOS los resultados directamente en
        # un DataFrame de pandas (una tabla en memoria).
        df = pd.read_sql_query(query, conn)
        print(f"Datos de la tabla '{tabla}' leídos correctamente. Se encontraron {len(df)} filas.")

        # --- 4. Procesamiento de Fechas (Timezone) ---
        # Excel no soporta fechas que incluyan información de zona horaria (timezone).
        # PostgreSQL a menudo guarda las fechas con esta información (ej. '...-05:00').
        # Este bucle revisa cada columna del DataFrame.
        for column in df.columns:
            # Si la columna es de tipo fecha y además tiene información de timezone...
            if pd.api.types.is_datetime64_any_dtype(df[column]) and df[column].dt.tz is not None:
                print(f"Convirtiendo la columna '{column}' a formato de fecha sin zona horaria.")
                # ...se elimina la información de timezone, dejando solo la fecha/hora.
                df[column] = df[column].dt.tz_localize(None)

        # --- 5. Escritura del Archivo Excel ---
        # Comprobamos si el archivo de Excel ya existe.
        file_exists = os.path.exists(archivo_excel)
        # Decidimos el modo de escritura: 'a' (append/añadir) si se quiere acumular
        # y el archivo ya existe. 'w' (write/escribir) en cualquier otro caso (crear de cero).
        mode = 'a' if acumular_si_existe and file_exists else 'w'
        
        # Se preparan los argumentos para pd.ExcelWriter. Hacemos esto en un diccionario
        # para poder añadir 'if_sheet_exists' solo cuando sea necesario.
        writer_args = {
            'engine': 'openpyxl',  # Le decimos a pandas que use openpyxl para poder formatear.
            'mode': mode
        }
        # El parámetro 'if_sheet_exists' solo funciona en modo 'a'. Le dice a pandas qué
        # hacer si la hoja que intentamos crear ya existe en el archivo.
        if mode == 'a':
            writer_args['if_sheet_exists'] = 'replace' # 'replace' borra la hoja vieja y pone la nueva.

        # 'with' asegura que el 'writer' de Excel se cierre correctamente al final.
        # Usamos **writer_args para pasar los argumentos del diccionario a la función.
        with pd.ExcelWriter(archivo_excel, **writer_args) as writer:
            # Escribimos el DataFrame a una hoja de Excel.
            # - sheet_name=tabla: nombra la hoja de Excel como la tabla de la BD.
            # - index=False: evita que pandas escriba el índice del DataFrame (0, 1, 2...) como una columna.
            df.to_excel(writer, sheet_name=tabla, index=False)

            # --- 6. Aplicación de Formato al Encabezado ---
            # Obtenemos el objeto 'worksheet' (la hoja de cálculo) desde el writer.
            worksheet = writer.sheets[tabla]
            # Creamos un objeto de estilo para el relleno de la celda.
            blue_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            # Creamos un objeto de estilo para la fuente (letra).
            white_font = Font(color="FFFFFF", bold=True)
            
            # worksheet[1] devuelve todas las celdas de la primera fila (el encabezado).
            # Iteramos sobre cada celda del encabezado para aplicarle los estilos.
            for cell in worksheet[1]:
                cell.fill = blue_fill
                cell.font = white_font
        
        # Se imprime un mensaje final para el usuario.
        if mode == 'a':
             print(f"Hoja '{tabla}' añadida/actualizada en el archivo existente '{archivo_excel}'.")
        else:
            if file_exists:
                print(f"Archivo '{archivo_excel}' sobrescrito con la nueva hoja '{tabla}'.")
            else:
                print(f"Nuevo archivo '{archivo_excel}' creado con la hoja '{tabla}'.")

    except (Exception, psycopg2.DatabaseError) as error:
        # Si algo falla en el bloque 'try', el código salta aquí.
        # Se captura el error y se muestra un mensaje descriptivo.
        print(f"Error durante la operación: {error}")
        
    finally:
        # Este bloque se ejecuta SIEMPRE, haya o no haya habido un error.
        # Es el lugar perfecto para asegurarse de cerrar la conexión.
        if conn: # Si la conexión se estableció con éxito...
            conn.close() # ...la cerramos.
            print("Conexión a la base de datos cerrada.")


def excel_to_postgres_inserts(excel_file, sheet_name, table_name, output_file, append=False):
    """
    Lee una hoja de cálculo de Excel y genera un archivo .sql con sentencias INSERT.
    
    Esta función es muy útil para preparar datos maestros o catálogos desde un
    Excel y luego poder cargarlos fácilmente en la base de datos.
    """
    # --- 1. Lectura del Archivo Excel ---
    # Se carga una hoja específica de un archivo Excel en un DataFrame de pandas.
    df = pd.read_excel(excel_file, sheet_name=sheet_name)
    
    # --- 2. Preparación del Archivo de Salida ---
    # Se decide el modo de apertura del archivo .sql: 'a' para añadir al final, 'w' para sobreescribir.
    mode = 'a' if append else 'w'
    
    # 'with' asegura que el archivo se cierre correctamente.
    # 'encoding='utf-8'' es importante para soportar caracteres especiales como acentos o 'ñ'.
    with open(output_file, mode, encoding='utf-8') as f:
        # --- 3. Iteración y Generación de SQL ---
        # df.iterrows() permite recorrer el DataFrame fila por fila.
        # Devuelve el índice de la fila y la fila misma como un objeto Series de pandas.
        for _, row in df.iterrows():
            # Elimina de la fila actual las columnas que no tienen valor (NaN).
            # Esto es útil para generar INSERTs solo con los campos que tienen datos.
            clean_row = row.dropna()
            
            # Si la fila no está completamente vacía...
            if not clean_row.empty:
                # Obtenemos los nombres de las columnas (que son los índices de la 'Series' clean_row).
                columns = ", ".join(f'"{col_name}"' for col_name in clean_row.index)
                
                # Procesamos los valores de cada celda para que sean compatibles con SQL.
                values = []
                for val in clean_row.values:
                    if isinstance(val, (int, float)):
                        # Los números se convierten a string directamente.
                        values.append(str(val))
                    else:
                        # Los textos (o fechas, etc.) necesitan un tratamiento especial:
                        # 1. Se convierten a string.
                        # 2. Se escapan las comillas simples (') duplicándolas ('') para evitar errores de SQL.
                        val_esc = str(val).replace("'", "''")
                        # 3. Se envuelven entre comillas simples para que PostgreSQL los reconozca como strings.
                        values.append(f"'{val_esc}'")
                
                # Se construye la sentencia INSERT final usando los nombres de columna y los valores procesados.
                insert_sql = f"INSERT INTO {table_name} ({columns}) VALUES ({', '.join(values)});"
                
                # Se escribe la sentencia en el archivo .sql, añadiendo un salto de línea.
                f.write(insert_sql + "\n")


def _split_sql_statements(sql_text: str) -> list:
    """Divide SQL en sentencias respetando dollar-quotes, strings y comentarios."""
    statements = []
    i = 0
    n = len(sql_text)
    start = 0

    while i < n:
        c = sql_text[i]

        # Comentario de línea: avanzar hasta fin de línea
        if c == '-' and i + 1 < n and sql_text[i + 1] == '-':
            while i < n and sql_text[i] != '\n':
                i += 1
            continue

        # Comentario multilínea /* ... */
        if c == '/' and i + 1 < n and sql_text[i + 1] == '*':
            i += 2
            while i < n - 1:
                if sql_text[i] == '*' and sql_text[i + 1] == '/':
                    i += 2
                    break
                i += 1
            continue

        # String con comilla simple: 'texto' (soporta '' como escape)
        if c == "'":
            i += 1
            while i < n:
                if sql_text[i] == "'" and i + 1 < n and sql_text[i + 1] == "'":
                    i += 2
                elif sql_text[i] == "'":
                    i += 1
                    break
                else:
                    i += 1
            continue

        # Dollar-quote: $tag$...$tag$ (incluye $$ sin tag)
        if c == '$':
            j = i + 1
            while j < n and sql_text[j] != '$' and sql_text[j] not in (' ', '\t', '\n', '\r'):
                j += 1
            if j < n and sql_text[j] == '$':
                tag = sql_text[i:j + 1]
                close = sql_text.find(tag, j + 1)
                if close != -1:
                    i = close + len(tag)
                    continue

        # Fin de sentencia
        if c == ';':
            stmt = sql_text[start:i].strip()
            if stmt and not all(
                not line.strip() or line.strip().startswith('--')
                for line in stmt.splitlines()
            ):
                statements.append(stmt)
            i += 1
            start = i
            continue

        i += 1

    # Texto restante sin ';' final
    stmt = sql_text[start:].strip()
    if stmt and not all(
        not line.strip() or line.strip().startswith('--')
        for line in stmt.splitlines()
    ):
        statements.append(stmt)

    return statements


def execute_sql_from_file(
    db_name, user, password, host, port, sql_file,
    log_file="", allow_partial=False
):
    """
    Se conecta a una BD PostgreSQL y ejecuta un script SQL instrucción por instrucción.

    log_file (str): Ruta completa del archivo de log. "" = sin log.
    allow_partial (bool): True = continuar aunque falle alguna instrucción;
                          False (defecto) = rollback completo al primer error.
    Retorna dict: success, total_stmts, ok_stmts, failed_stmts, errors.
    """
    result = {
        "file": sql_file,
        "success": False,
        "total_stmts": 0,
        "ok_stmts": 0,
        "failed_stmts": 0,
        "errors": [],
    }

    if not os.path.exists(sql_file):
        result["errors"].append("Archivo no encontrado en disco")
        _write_log(log_file, sql_file, None, "ARCHIVO NO ENCONTRADO")
        return result

    with open(sql_file, "r", encoding="utf-8") as f:
        contenido = f.read()

    stmts = _split_sql_statements(contenido)

    if not stmts:
        result["success"] = True
        _write_log(log_file, sql_file, None, "ARCHIVO VACIO - sin instrucciones ejecutables")
        return result

    conn = None
    try:
        conn = psycopg2.connect(
            dbname=db_name, user=user, password=password, host=host, port=port
        )
        _write_log(log_file, sql_file, None, f"INICIO - {len(stmts)} instrucciones")

        for i, stmt in enumerate(stmts, start=1):
            result["total_stmts"] += 1
            try:
                with conn.cursor() as cur:
                    cur.execute(stmt)
                result["ok_stmts"] += 1
                _write_log(log_file, sql_file, i, "OK")
            except Exception as err:
                result["failed_stmts"] += 1
                result["errors"].append(f"STMT {i}: {err}")
                _write_log(log_file, sql_file, i, f"ERROR: {err}")
                if not allow_partial:
                    conn.rollback()
                    _write_log(log_file, sql_file, None, "ROLLBACK COMPLETO")
                    return result

        conn.commit()
        result["success"] = result["failed_stmts"] == 0
        estado = "EXITO COMPLETO" if result["success"] else f"PARCIAL ({result['failed_stmts']} errores)"
        _write_log(log_file, sql_file, None, f"FIN - {estado}")

    except Exception as error:
        result["errors"].append(f"Error de conexion: {error}")
        _write_log(log_file, sql_file, None, f"ERROR DE CONEXION: {error}")
        if conn:
            conn.rollback()
    finally:
        if conn:
            conn.close()

    return result


def _write_log(log_file: str, sql_file: str, stmt_num, mensaje: str):
    """Escribe una línea de log si log_file es no vacío."""
    if not log_file:
        return
    import datetime
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    archivo = os.path.basename(sql_file) if sql_file else ""
    stmt_part = f"[STMT {stmt_num}] " if stmt_num is not None else ""
    linea = f"[{ts}] [{archivo}] {stmt_part}{mensaje}\n"
    try:
        with open(log_file, "a", encoding="utf-8") as f:
            f.write(linea)
    except Exception:
        pass
